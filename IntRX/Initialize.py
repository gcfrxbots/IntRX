import os
import argparse
import time
import shutil
import json
from json import JSONDecodeError
from websocket import create_connection, WebSocketConnectionClosedException, WebSocketBadStatusException
import subprocess
import datetime
from distutils.dir_util import copy_tree

try:
    import xlrd
    import xlsxwriter
    from openpyxl import load_workbook
    import pyperclip
except ImportError as e:
    print(e)
    raise ImportError(">>> One or more required packages are not properly installed! Run INSTALL_REQUIREMENTS.bat to fix!")


parser = argparse.ArgumentParser(description='Generate Settings File')
parser.add_argument('--g', dest="GenSettings", action="store_true")
parser.add_argument('--d', dest="debugMode", action="store_true")
parser.set_defaults(GenSettings=False, debugMode=False)

GenSettings = (vars(parser.parse_args())["GenSettings"])
debugMode = (vars(parser.parse_args())["debugMode"])

'''----------------------SETTINGS----------------------'''

'''FORMAT ---->   ("Option", "Default", "This is a description"), '''
defaultSettings = [
    ("PLATFORM", "Twitch", "Choose your streaming platform: Twitch, Youtube, Caffeine, Brime, Trovo, Glimesh, Dlive."),
    ("CHAT AS RXBOTS", "No", "Set to Yes for all bot messages in your chat to be sent from the user rxbots, or No if you want the messages to be from your own account or your own bot account. (Yes/No)"),
    ("", "", ""),
    ("PREFIX", "!", "The symbol or prefix to tell the bot that a message is a command intended for it. Set to No for no prefix at all (For twitch plays or similar). Recommend ! ? $ or #."),
    ("ANNOUNCE GAME", "Yes", "Announce in chat when you begin playing a game that the bot supports. (Yes/No)"),
    ("CD BETWEEN CMDS", 15, "The cooldown, in seconds, between two consecutive commands."),
    ("DEFAULT ARG", "No", "Set to No to require an argument for any commands with %ARGS%. Set to anything else to use that value if the user doesn't provide an argument. "),
    ("MAX ARG", 25, "The highest integer that a user can provide for %ARGS%. Used to prevent things like $SPAM from being run 99999 times."),
    ("", "", ">>>COMMAND PHRASES BELOW - Do not touch these settings unless you want IntRX to be triggered by a different bot."),
    ("ALT BOT NAMES", "", "If you use another bot which uses its own account, such as NightBot or StreamElements, type their names here. Separate by a comma."),
    ("COMMAND PHRASE 1", "", "If this isn't empty, commands will only be accepted from your bot(s), and must be executed via this phrase. Write any phrase that includes %CMD%, which marks where the bot will look for a command. Ex: 'just executed the %CMD% command'. Check the documentation for more info."),
    ("COMMAND PHRASE 2", "", "Second command phrase the bot will search for."),
    ("COMMAND PHRASE 3", "", "Third command phrase the bot will search for."),
]
'''----------------------END SETTINGS----------------------'''

# TODO - Fix non-global sub/dono and all rewards for cmds - Done?
# TODO - Make CHAT AS RXBOTS work - See message from END3R - Done?
# TODO - Mouse Movement / Control
# TODO - Pipeline for compiling
# TODO - Documentation, ew


def stopBot(err):
    print(">>>>>---------------------------------------------------------------------------<<<<<")
    print(err)
    print("More info can be found here: https://rxbots.net/intrx-setup.html\n")
    print(">>>>>----------------------------------------------------------------------------<<<<<")
    time.sleep(3)
    quit()


def formatSettingsXlsx():
    try:
        with xlsxwriter.Workbook('../Config/Settings.xlsx') as workbook:  # FORMATTING
            worksheet = workbook.add_worksheet('Settings')
            format = workbook.add_format({'bold': True, 'center_across': True, 'font_color': 'white', 'bg_color': 'gray'})
            boldformat = workbook.add_format({'bold': True, 'center_across': True, 'font_color': 'white', 'bg_color': 'black'})
            lightformat = workbook.add_format({'bold': True, 'center_across': True, 'font_color': 'black', 'bg_color': '#DCDCDC', 'border': True})
            worksheet.set_column(0, 0, 25)
            worksheet.set_column(1, 1, 50)
            worksheet.set_column(2, 2, 180)
            worksheet.write(0, 0, "Option", format)
            worksheet.write(0, 1, "Your Setting", boldformat)
            worksheet.write(0, 2, "Description", format)
            worksheet.set_column('B:B', 50, lightformat)  # END FORMATTING

            row = 1  # WRITE SETTINGS
            col = 0
            for option, default, description in defaultSettings:
                worksheet.write(row,  col, option)
                if isinstance(default, int):
                    worksheet.write(row,  col + 1, default)
                else:
                    worksheet.write(row, col + 1, str(default))  # This prevents Yes and No from being written as bools in excel
                worksheet.write(row,  col + 2, description)
                row += 1
        print("Settings.xlsx has been configured successfully.")
    except PermissionError:
        stopBot("Can't open the settings file. Please close it and make sure it's not set to Read Only")


def formatInteractxlsx():
    try:
        with xlsxwriter.Workbook('../Config/InteractConfig.xlsx') as workbook:  # FORMATTING
            listGames = ("Skyrim", "Oblivion", "Fallout 4", "Fallout NV", "Fallout 3", "Minecraft", "Subnautica", "Witcher 3", "Valheim")

            format = workbook.add_format({'bold': True, 'center_across': True, 'font_color': 'white', 'bg_color': 'gray'})
            lightformat = workbook.add_format({'center_across': True, 'font_color': 'black', 'bg_color': '#DCDCDC', 'border': True})
            evenlighterformat =  workbook.add_format({'center_across': True, 'font_color': 'black', 'bg_color': '#f0f0f0', 'border': True})
            redformat = workbook.add_format({'font_color': 'black', 'bg_color': '#ffdede', 'border': True})
            greenformat = workbook.add_format({'font_color': 'black', 'bg_color': '#e6ffd4', 'border': True})

            worksheet = workbook.add_worksheet("Global")  # FORMAT GLOBAL
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 10)
            worksheet.set_column(2, 2, 10)
            worksheet.set_column(3, 3, 45)
            worksheet.set_column(4, 4, 100)
            worksheet.set_column(5, 5, 20)
            worksheet.write(0, 0, "Command", format)
            worksheet.write(0, 1, "Cooldown", format)
            worksheet.write(0, 2, "Disable", format)
            worksheet.write(0, 3, "Active Window", format)
            worksheet.write(0, 4, "What to Run", format)
            worksheet.write(0, 5, "Sub Only", format)
            worksheet.write(0, 6, "Donation Cost", format)
            worksheet.write(0, 7, "Reward/Min. Cost", format)
            worksheet.set_column('B:B', 10, lightformat)
            worksheet.set_column('C:C', 10, redformat)
            worksheet.set_column('D:D', 30, evenlighterformat)
            worksheet.set_column('F:H', 20, greenformat)

            # COMMENTS FOR ADDITIONAL CLARITY

            worksheet.write_comment('A1', "This is what people will type in chat to run your command. If you don't specify a prefix, the bot will automatically add the PREFIX setting to the start (Like dragon will be !dragon unless you have no prefix)")
            worksheet.write_comment('B1', "This is how many seconds must pass after this command is run before any user can run it again.")
            worksheet.write_comment('C1', "Leave empty unless you want to temporarily disable the command, then just type Yes in the field.")
            worksheet.write_comment('D1', "The name of the window that must be active for this command to work. To get a window name, hover over the icon in your taskbar and the text at the top of the popup is your active window. You can shorten it as well (Like Minecraft will work for windows named Minecraft 1.12 or Minecraft 1.19)")
            worksheet.write_comment('E1', "Specify an executable located in the UserScripts folder, or any of the following: $PRESS [key] | $HOLD [key] [seconds] | $SPAM [key] [times] | $TYPE [message] | $CHAT [message] | $RUN [file in UserScripts] | $WAIT [seconds]")
            worksheet.write_comment('F1', "Set to Yes if only subscribers can run this, or No (or leave empty) if anyone can.")
            worksheet.write_comment('G1', "Specify a donation value that must be included with the command for the command to run. This only works with built-in chat donations like Bits on Twitch.")
            worksheet.write_comment('H1', "Choose the name of a reward that must be redeemed with the command for the command to run. Alternately, if you provide a number, any channel reward with a point cost greater than or equal to that number will work.")

            for item in listGames:  # FORMAT GAMES
                worksheet = workbook.add_worksheet(item)
                worksheet.set_column(0, 0, 30)
                worksheet.set_column(1, 1, 10)
                worksheet.set_column(2, 2, 10)
                worksheet.set_column(3, 3, 110)
                worksheet.set_column(4, 4, 20)
                worksheet.write(0, 0, "Command", format)
                worksheet.write(0, 1, "Cooldown", format)
                worksheet.write(0, 2, "Disable", format)
                worksheet.write(0, 3, "Command To Execute", format)
                worksheet.write(0, 4, "Sub Only", format)
                worksheet.write(0, 5, "Donation Cost", format)
                worksheet.write(0, 6, "Reward to Redeem", format)
                worksheet.set_column('B:B', 10, lightformat)  # END FORMATTING
                worksheet.set_column('C:C', 10, redformat)  # END FORMATTING
                worksheet.set_column('E:G', 20, greenformat)

                # COMMENTS FOR ADDITIONAL CLARITY

                worksheet.write_comment('A1', "THIS COMMAND WILL ONLY WORK WHEN THE GAME IS ACTIVE! This is what people will type in chat to run your command. If you don't specify a prefix, the bot will automatically add the PREFIX setting to the start (Like dragon will be !dragon unless you have no prefix)")
                worksheet.write_comment('B1', "This is how many seconds must pass after this command is run before any user can run it again.")
                worksheet.write_comment('C1', "Leave empty unless you want to temporarily disable the command, then just type Yes in the field.")
                worksheet.write_comment('D1', "The in-game console command the bot should run when the command is executed. Make sure the command is a valid command, and your game's console is enabled!")
                worksheet.write_comment('E1', "Set to Yes if only subscribers can run this, or No (or leave empty) if anyone can.")
                worksheet.write_comment('F1', "Specify a donation value that must be included with the command for the command to run. This only works with built-in chat donations like Bits on Twitch.")
                worksheet.write_comment('G1', "Choose the name of a reward that must be redeemed with the command for the command to run. Alternately, if you provide a number, any channel reward with a point cost greater than or equal to that number will work.")


        print("InteractConfig.xlsx has been updated successfully.")
    except PermissionError:
        stopBot("Can't open the InteractConfig file. Please close it and make sure it's not set to Read Only")


def readSettings():
    wb = xlrd.open_workbook('../Config/Settings.xlsx')
    sheet = wb.sheet_by_name("Settings")
    for item in range(sheet.nrows):
        if item == 0:
            pass
        else:
            option = sheet.cell_value(item,0)
            setting = sheet.cell_value(item,1)

            if setting == "No" and option == "PREFIX":  # Exception for PREFIX to make the option an empty string rather than a bool
                setting = ""

            if setting == "Yes":
                setting = True
            if setting == "No":
                setting = False


            settings[option] = setting

    # Check for new/changed settings
    if sheet.nrows != (len(defaultSettings)+1):
        for item in settings:
            for i in enumerate(defaultSettings):
                if (i[1][0]) == item:
                    defaultSettings[i[0]] = (item, settings[item], defaultSettings[i[0]][2])
        formatSettingsXlsx()
        stopBot("The settings for IntRX have changed since you last started the script. Settings.xlsx has updated, go check it out.")

    return settings


def changeChatSetting(setting):  # An overcomplicated function that exists only to allow users to be lazy and not have to change the CHAT AS RXBOTS setting themselves.
    try:
        workbook = load_workbook('../Config/Settings.xlsx')
        ws = workbook["Settings"]
        cellToEdit = None

        for settingEntry in enumerate(defaultSettings):
            row = settingEntry[0] + 2
            if settingEntry[1][0] == "CHAT AS RXBOTS":
                cellToEdit = "B" + str(row)

        if setting:
            ws[cellToEdit] = "Yes"
        else:
            ws[cellToEdit] = "No"

        workbook.save('../Config/Settings.xlsx')

    except PermissionError:
        stopBot("Can't read/edit Settings.xlsx! Please close it and make sure it isn't read only!")




def getPlatform():
    validPlatforms = ["twitch", "youtube", "trovo", "glimesh", "brime", "caffeine", "dlive"]
    wb = xlrd.open_workbook('../Config/Settings.xlsx')
    sheet = wb.sheet_by_name("Settings")
    for item in range(sheet.nrows):
        if item == 0:
            pass
        else:
            option = sheet.cell_value(item,0)
            setting = sheet.cell_value(item,1)
            if option == "PLATFORM":
                if setting:
                    if setting.lower() in validPlatforms:
                        return setting.strip().lower()
                    else:
                        print("Platform %s is invalid! Please change the PLATFORM setting to a valid supported platform." % setting )
                else:
                    print("No platform detected.")


def authenticate():
    path = os.path.abspath("../Setup/Authenticate.bat")
    subprocess.call(path)


def initSetup():
    global settings
    settings = {}
    killbot = False
    if not os.path.exists('../Config'):
        os.mkdir("../Config")

    if not os.path.exists('../Config/tokens'):
        os.mkdir("../Config/tokens")

    if not os.path.exists('../Config/UserScripts'):
        os.mkdir("../Config/UserScripts")

        for file in os.listdir("./Resources/Included Scripts"):
            full_file_name = os.path.join("./Resources/Included Scripts", file)
            if os.path.isfile(full_file_name) and (".ahk" not in full_file_name):
                shutil.copy(full_file_name, "../Config/UserScripts")


    if not os.path.exists("../Config/UserScripts/Templates"):
        copy_tree("./Resources/Templates", "../Config/UserScripts/Templates")

    if not os.path.exists('../Config/Settings.xlsx'):
        formatSettingsXlsx()
        killbot = True

    if not os.path.exists('../Config/InteractConfig.xlsx'):
        print("Creating InteractionConfig.xlsx")
        killbot = True
        formatInteractxlsx()

    if killbot:
        stopBot("\nPlease open the Config folder and edit Settings.xlsx by following the readme, then start the bot again. \nYou should also set the platform you would like to connect to.")

    # Read the settings file
    settings = readSettings()

    if settings["COMMAND PHRASE 1"] or settings["COMMAND PHRASE 2"] or settings["COMMAND PHRASE 3"] :
        settingsCmdPhrases = [settings["COMMAND PHRASE 1"].replace("%CMD%", "%cmd%"), settings["COMMAND PHRASE 2"].replace("%CMD%", "%cmd%"), settings["COMMAND PHRASE 3"].replace("%CMD%", "%cmd%")]
        for phrase in settingsCmdPhrases:
            if phrase:
                if not "%cmd%" in phrase:
                    stopBot("One of your COMMAND PHRASE settings does not have %CMD% in it anywhere.")
                if len(phrase.strip()) < 10:
                    stopBot("One of your COMMAND PHRASE settings is too short, it needs to be at least 10 characters long (including spaces and %CMD%) so the bot can detect the actual phrase.")

        print("\n\n IMPORTANT! You have a COMMAND PHRASE set in your Settings. NORMAL COMMANDS WON'T WORK!"
              "\n The bot will ONLY run commands via phrases sent only by the specified bot accounts. "
              "\n If you don't know what this means, remove your COMMAND PHRASE settings and read the documentation.\n")
        time.sleep(2)

    # Set the normal token
    platform = getPlatform()

    if os.path.exists("../Config/tokens/token_{platform}.txt".format(platform=platform)):
        with open("../Config/tokens/token_{platform}.txt".format(platform=platform), "r") as f:
            chatConnection.token = f.read()
            f.close()
    else:
        print("Casterlabs authentication needed...")
        time.sleep(1)
        # Run authenticate separately to avoid import loops
        authenticate()

        time.sleep(1)
        with open("../Config/tokens/token_{platform}.txt".format(platform=platform), "r") as f:
            chatConnection.token = f.read()
            f.close()

    # Set the puppet token, if it exists
    if os.path.exists("../Config/tokens/puppet_{platform}.txt".format(platform=platform)):
        chatConnection.puppet = True
        with open("../Config/tokens/puppet_{platform}.txt".format(platform=platform), "r") as f:
            chatConnection.puppetToken = f.read()
            f.close()


    # Read settings again after them potentially being changed
    settings = readSettings()

    print(">> Initial checkup complete!")
    return settings


class chat:
    global settings

    def __init__(self):
        self.ws = None
        self.url = "wss://api.casterlabs.co/v2/koi?client_id=jJu2vQGnHf5U5trv"
        self.caffeineUrl = "wss://api.casterlabs.co/v2/koi?client_id=LmHG2ux992BxqQ7w9RJrfhkW"
        self.puppet = False
        self.active = False
        self.token = None
        self.puppetToken = None

    def login(self):
        loginRequest = {
                "type": "LOGIN",
                "token": self.token
            }
        self.ws.send(json.dumps(loginRequest))
        time.sleep(1)

    def puppetlogin(self):
        time.sleep(1.5)
        loginRequest = {
            "type": "PUPPET_LOGIN",
            "token": self.puppetToken
        }
        self.ws.send(json.dumps(loginRequest))

    def sendRequest(self, request):
        self.ws.send(json.dumps(request))

    def sendToChat(self, message):
        if message:
            request = {
                "type": "CHAT",
                "message": message,
                "chatter": "CLIENT",
                "isUserGesture": False,
                "reply_target": None}

            if self.puppet:
                request['chatter'] = "PUPPET"

            if settings["CHAT AS RXBOTS"]:
                request["chatter"] = "SYSTEM"

            self.sendRequest(request)

    def start(self, silent=False, reconnect=False):
        try:
            if getPlatform() == "caffeine":
                self.ws = create_connection(self.caffeineUrl)
            else:
                self.ws = create_connection(self.url)
            self.login()
            if not silent:
                print(">> Connecting to platform: " + getPlatform().capitalize())
        except (WebSocketConnectionClosedException, WebSocketBadStatusException):
            if not reconnect:  # Only tries to reconnect if its not already being called from reconnect to prevent loops
                time.sleep(5)
                self.reconnect()
            pass


    def reconnect(self):
        for x in range(10):
            try:
                self.ws = None
                self.start(silent=True, reconnect=True)
                return
            except:
                time.sleep(5)
        raise ConnectionError("Unable to connect to Koi after multiple reconnect attempts. Please wait a few minutes then restart the bot.")


class runMiscControls:

    def __init__(self):
        self.timerActive = False
        self.timers = {}

    def getUser(self, line):
        seperate = line.split(":", 2)
        user = seperate[1].split("!", 1)[0]
        return user

    def getMessage(self, line):
        seperate = line.split(":", 2)
        message = seperate[2]
        return message

    def formatTime(self):
        return datetime.datetime.today().now().strftime("%I:%M")

    def setTimer(self, name, duration):
        self.timerActive = True
        curTime = datetime.datetime.now()
        targetTime = curTime + datetime.timedelta(seconds=duration)
        self.timers[name] = targetTime

    def timerDone(self, timer):
        self.timers.pop(timer)
        print(timer + " timer complete.")
        if not self.timers:
            self.timerActive = False


misc = runMiscControls()
chatConnection = chat()


if GenSettings:
    if not os.path.exists('../Config'):
        os.mkdir("../Config")

    if not os.path.exists('../Config/Settings.xlsx'):
        initSetup()
        print("\nPlease open Config / Settings.xlsx and configure the bot, then run it again.")
        print("Please read the readme to get everything set up!")
        time.sleep(3)
        quit()
    else:
        print("Everything is already set up!")

